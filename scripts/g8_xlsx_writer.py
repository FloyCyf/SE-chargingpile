#!/usr/bin/env python3
"""
G8 修正版 xlsx 生成器
==========================
读取原 G8 用例 xlsx 的事件序列, 用 g8_simulator 的精确模拟结果
重新填写"需填写sheet"和"账单和详单明细"两个工作表, 输出修正版.

运行:
  .venv/Scripts/python.exe scripts/g8_xlsx_writer.py
"""
from __future__ import annotations
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 引入 simulator
sys.path.insert(0, str(Path(__file__).resolve().parent))
from g8_simulator import (Simulator, EVENTS, t, fmt_time, calc_bill,
                          BASE_DATE, PILE_QUEUE_LEN)


# ---------------------------------------------------------------------------
#  时刻列表(事件时刻 + 完成时刻 + 故障恢复时刻)
# ---------------------------------------------------------------------------
def run_simulation_with_snapshots() -> tuple[Simulator, list[dict]]:
    """运行模拟, 在每个事件时刻 + 每次完成/恢复时刻拍快照"""
    sim = Simulator()
    snapshots = []

    def snap(label: str, event_str: str):
        snapshots.append({
            "vtime": fmt_time(sim.now),
            "label": label,
            "event": event_str,
            "snap": sim.snapshot(),
        })

    # Hook: 给 sim 注入"完成时也拍快照"的能力
    orig_finish = sim.finish_charging

    def wrapped_finish(pile, status):
        r = orig_finish(pile, status)
        if r is not None:
            # 用 'completion' 标记, 后续 writer 决定是否独立成行
            snap("completion", f"{pile.pid}: 完成一辆车")
        return r
    sim.finish_charging = wrapped_finish  # type: ignore

    orig_recover = sim.recover_pile

    def wrapped_recover(pid):
        orig_recover(pid)
        snap("recovery", f"{pid}: 故障恢复")
    sim.recover_pile = wrapped_recover  # type: ignore

    # 主循环: 每个事件执行
    for vtime, etype, target, ctype, value in EVENTS:
        sim.advance_to(t(vtime))
        if etype == "submit":
            r = sim.submit(target, ctype, value)
            ev = f"(A,{target},{ctype},{value})"
        elif etype == "cancel":
            r = sim.cancel(target)
            ev = f"(A,{target},O,0)"
        elif etype == "modify":
            r = sim.modify(target, ctype, value)
            ev = f"(C,{target},{ctype},{value})"
        elif etype == "fault":
            r = sim.fault(target, value)
            ev = f"(B,{target},O,{int(value)})"
        else:
            ev = "?"
        snap("event", ev)
        sim.dispatch_waiting()

    # 最后推进到所有车结束
    sim.advance_to(t("16:00"))
    snap("end", "全部结束")
    return sim, snapshots


# ---------------------------------------------------------------------------
#  Sheet 1: 需填写sheet (157 行 x 10 列, 3 行/事件)
# ---------------------------------------------------------------------------
COL_TIME = 1
COL_EVENT = 2
COL_F1 = 3   # 快充1
COL_F2 = 4
COL_F3 = 5
COL_T1 = 6   # 慢充1
COL_T2 = 7
COL_WAITING = 8
COL_FAULT_NOTE = 9


def build_pile_cell_value(pile_data: dict, pos: int) -> str:
    """构造单个桩的某位置(0/1/2)的 cell 文本.
       格式: (车号,已充电量,当前费用)  或  空字符串  或  '故障中'"""
    if pile_data["status"] == "FAULT":
        return "故障中" if pos == 0 else ""
    q = pile_data["queue"]
    if pos >= len(q):
        return ""
    c = q[pos]
    return f"({c['vid']},{c['charged_kwh']:.2f},{c['current_fee']:.2f})"


def build_waiting_cell(snap: dict) -> str:
    """等候区 H 列: 多行, 每行 (车号,类型,充电量)"""
    lines = []
    for c in snap["waiting_fast"] + snap["waiting_slow"]:
        lines.append(f"({c['vid']},{c['ctype']},{c['req_kwh']:.2f})")
    return "\n".join(lines)


def build_fault_note(snap: dict, event_str: str) -> str:
    """I 列: 故障后转移记录 + 备注"""
    notes = []
    if snap["fault_fast"]:
        ids = "/".join(c["vid"] for c in snap["fault_fast"])
        notes.append(f"快充故障队列: {ids}")
    if snap["fault_slow"]:
        ids = "/".join(c["vid"] for c in snap["fault_slow"])
        notes.append(f"慢充故障队列: {ids}")
    return "\n".join(notes)


def write_state_sheet(ws, snapshots: list[dict]):
    """写"需填写sheet"."""
    thin = Side(border_style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="FFF2CC")

    # Header
    ws.cell(1, 1, "")
    ws.cell(1, 3, "(车号,已充电量,当前费用)")
    ws.merge_cells(start_row=1, end_row=1, start_column=3, end_column=7)
    ws.cell(1, 8, "(车号,充电类型,充电量)")

    headers2 = ["时间", "事件", "快充1", "快充2", "快充3", "慢充1", "慢充2",
                "等候区(10车位)", "故障后转移记录(说明)"]
    for c, h in enumerate(headers2, 1):
        cell = ws.cell(2, c, h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border

    # 只把"event"类型的快照渲染为行(每个事件占 3 行: 桩 pos 0/1/2)
    event_snaps = [s for s in snapshots if s["label"] == "event"]
    # 加上完成/恢复的合并(显示在下一个事件前的"结果"附注)

    current_row = 3
    for ev in event_snaps:
        snap = ev["snap"]
        vtime = ev["vtime"] + ":00"
        event_str = ev["event"]

        # 3 行: pos 0, 1, 2
        for pos in range(PILE_QUEUE_LEN):
            row = current_row + pos
            if pos == 0:
                ws.cell(row, COL_TIME, vtime).font = Font(bold=True)
                ws.cell(row, COL_EVENT, event_str)
                ws.cell(row, COL_WAITING, build_waiting_cell(snap))
                fnote = build_fault_note(snap, event_str)
                if fnote:
                    ws.cell(row, COL_FAULT_NOTE, fnote)
            # 桩
            ws.cell(row, COL_F1,
                    build_pile_cell_value(snap["piles"]["F1"], pos))
            ws.cell(row, COL_F2,
                    build_pile_cell_value(snap["piles"]["F2"], pos))
            ws.cell(row, COL_F3,
                    build_pile_cell_value(snap["piles"]["F3"], pos))
            ws.cell(row, COL_T1,
                    build_pile_cell_value(snap["piles"]["T1"], pos))
            ws.cell(row, COL_T2,
                    build_pile_cell_value(snap["piles"]["T2"], pos))
            for c in range(1, 10):
                ws.cell(row, c).border = border
                ws.cell(row, c).alignment = Alignment(
                    vertical="center", wrap_text=True)
        current_row += PILE_QUEUE_LEN

    # 追加: 完成/恢复事件的"观察点"行
    obs_snaps = [s for s in snapshots
                 if s["label"] in ("completion", "recovery")]
    if obs_snaps:
        # 合并相同 vtime 的观察事件
        merged: dict[str, list[str]] = {}
        for s in obs_snaps:
            merged.setdefault(s["vtime"], []).append(s["event"])
        for vtime in sorted(merged.keys()):
            evs = merged[vtime]
            ws.cell(current_row, COL_TIME, vtime + ":00")
            ws.cell(current_row, COL_EVENT,
                    " / ".join(evs)).fill = section_fill
            for c in range(1, 10):
                ws.cell(current_row, c).border = border
                ws.cell(current_row, c).alignment = Alignment(
                    vertical="center", wrap_text=True)
            current_row += 1

    # 列宽
    widths = [10, 16, 22, 22, 22, 22, 22, 22, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 行高: 数据行(3 行/事件)用 60
    for r in range(3, current_row):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 36


# ---------------------------------------------------------------------------
#  Sheet 2: 说明
# ---------------------------------------------------------------------------
def write_readme_sheet(ws):
    rows = [
        ["G8 测试用例(修正版)说明", "", ""],
        ["", "", ""],
        ["项目", "值", "说明"],
        ["时间比例", "1:10",
         "真实 30 分钟 = 虚拟 300 分钟, 输入窗口 06:00-11:00"],
        ["充电桩配置", "3 快 + 2 慢, 每桩 3 车位, 等候区 10 车位",
         "快充 30 kW, 慢充 10 kW"],
        ["计费", "峰 1.0 / 平 0.7 / 谷 0.4, 服务费 0.8 元/度",
         "峰: 10-15 / 18-21; 平: 7-10 / 15-18 / 21-23; 谷: 其余"],
        ["调度策略", "正常: 同模式 FIFO 叫号 + 最短(等待+自己充电)选桩\n"
                     "故障: 优先级调度(A 策略)\n"
                     "故障恢复: 时间顺序重排",
         "见详细需求 §3-§7"],
        ["", "", ""],
        ["与原 G8 用例的差异", "", ""],
        ["1. V21 事件 (10:10)",
         "原: (C,V21,F,10) 修改 → 系统必拒绝(规则禁止充电区修改)\n"
         "改: (A,V21,O,0) 取消, V21 已充 10 度按取消结算 16.50 元",
         "对齐 g8_test.py 现行变通"],
        ["2. 故障 A 策略时机",
         "原: T1 故障(10:30)后 V10/V16 滞留, 直到 13:40 才进 T2 充电\n"
         "改: 规则要求'其它桩队列有空位时立即调度故障队列', 故 V10 在\n"
         "    11:25 V12 完成时即进 T2 pos2; T1 11:30 恢复时再按时间顺\n"
         "    序重排, V16 → T1 pos0(11:30 开始充), V10 移到 T1 pos1",
         "见详细需求 §7.a + §7.c"],
        ["3. V19 状态",
         "原: g8_test.py 写 V19 FAULTED 4kWh(把 V19 当成在 F1 充电)\n"
         "改: V19 在 10:05 改为快充后实际进 F3 pos1, F1 故障与之无关,\n"
         "    V19 在 10:40 V14 完成后开始充电, 11:30 完成 25 kWh = 45 元",
         "F1 故障 ≠ F3 故障"],
        ["4. 时间精度", "完成时刻已四舍五入到分钟", "原表也是分钟粒度"],
    ]
    for r, row_data in enumerate(rows, 1):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            elif r == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
            elif r == 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40
    for r in range(4, len(rows) + 1):
        ws.row_dimensions[r].height = 60


# ---------------------------------------------------------------------------
#  Sheet 3: 账单和详单明细
# ---------------------------------------------------------------------------
def write_bills_sheet(ws, sim: Simulator):
    headers = [
        "车辆ID", "结算状态", "未生成账单原因", "充电类型", "充电桩",
        "开始时间", "结束/状态时间", "请求电量(度)", "实际充电(度)",
        "谷时电量(度)", "谷时费(元)", "平时电量(度)", "平时费(元)",
        "峰时电量(度)", "峰时费(元)", "服务费(元)", "总费用(元)",
        "充电时长(分钟)", "分时段详单/备注",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    status_zh = {
        "COMPLETED": "已完成",
        "CANCELLED": "已取消",
        "FAULTED": "故障中断-部分结算",
    }
    row = 2
    for vid in sorted(sim.cars.keys(), key=lambda v: int(v[1:])):
        c = sim.cars[vid]
        bill = getattr(c, "bill", None)
        st = c.status
        st_zh = status_zh.get(st, st)
        if st == "CANCELLED" and c.charged_kwh > 0:
            st_zh = "已取消-部分结算"
        no_bill_reason = ""
        if st == "CANCELLED" and (bill is None or c.charged_kwh == 0):
            no_bill_reason = (f"{fmt_time(c.finish_time) or '?'}"
                              f" 取消时未开始充电")
            st_zh = "已取消-未充电"

        ws.cell(row, 1, vid)
        ws.cell(row, 2, st_zh)
        ws.cell(row, 3, no_bill_reason)
        ws.cell(row, 4, "快充" if c.charge_type == "Fast" else "慢充")
        ws.cell(row, 5, c.final_pile or "")
        ws.cell(row, 6, fmt_time(c.charge_start) or "")
        ws.cell(row, 7, fmt_time(c.finish_time) or "")
        ws.cell(row, 8, c.requested_kwh)
        if bill:
            ws.cell(row, 9, round(c.charged_kwh, 2))
            ws.cell(row, 10, bill.get("valley_kwh", 0))
            ws.cell(row, 11, bill.get("valley_fee", 0))
            ws.cell(row, 12, bill.get("flat_kwh", 0))
            ws.cell(row, 13, bill.get("flat_fee", 0))
            ws.cell(row, 14, bill.get("peak_kwh", 0))
            ws.cell(row, 15, bill.get("peak_fee", 0))
            ws.cell(row, 16, bill.get("service_fee", 0))
            ws.cell(row, 17, bill.get("total_fee", 0))
            ws.cell(row, 18, bill.get("duration_min", 0))
            seg_lines = []
            period_zh = {"peak": "峰时", "flat": "平时", "valley": "谷时"}
            for seg in bill.get("segments", []):
                seg_lines.append(
                    f"{period_zh.get(seg['period'], seg['period'])} "
                    f"{seg['start']}-{seg['end']} "
                    f"{seg['minutes']}分钟 "
                    f"{seg['kwh']:.2f}度*{seg['rate']}={seg['fee']:.2f}")
            ws.cell(row, 19, "\n".join(seg_lines))
        else:
            ws.cell(row, 9, 0)
            ws.cell(row, 19, no_bill_reason)
        # 样式
        for cc in range(1, 20):
            ws.cell(row, cc).alignment = Alignment(
                vertical="top", wrap_text=True)
        row += 1

    # 列宽
    widths = [8, 18, 28, 8, 8, 10, 10, 12, 12, 12, 12, 12, 12,
              12, 12, 10, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, row):
        ws.row_dimensions[r].height = 70
    ws.row_dimensions[1].height = 36


# ---------------------------------------------------------------------------
#  主入口
# ---------------------------------------------------------------------------
def main():
    print("[1/3] 运行 G8 修正版模拟...")
    sim, snapshots = run_simulation_with_snapshots()
    print(f"      事件数={len(EVENTS)} 总快照={len(snapshots)} "
          f"成车={len(sim.cars)}")

    print("[2/3] 构造工作簿...")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "需填写sheet"
    write_state_sheet(ws1, snapshots)

    ws2 = wb.create_sheet("说明")
    write_readme_sheet(ws2)

    ws3 = wb.create_sheet("账单和详单明细")
    write_bills_sheet(ws3, sim)

    out = Path(__file__).resolve().parent.parent / "docs" / \
        "G8测试用例_修正版.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    print(f"[3/3] 写入 {out}")
    wb.save(str(out))
    print(f"\n  完成: {out}")
    print(f"  Sheet1='需填写sheet' (3 行/事件 + 完成观察行)")
    print(f"  Sheet2='说明' (差异说明)")
    print(f"  Sheet3='账单和详单明细' (22 车结算)\n")


if __name__ == "__main__":
    main()
