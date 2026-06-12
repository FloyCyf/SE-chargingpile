from __future__ import annotations

import copy
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).resolve().parent))
from g8_xlsx_writer import run_simulation_with_snapshots  # noqa: E402
from g8_simulator import Simulator, EVENTS, t  # noqa: E402


ROOT = Path(__file__).resolve().parent.parent
SOURCE = Path(r"D:\NeatDownload\作业验收用例（包含参数说明）.xlsx")
OUTPUT = ROOT / "docs" / "G8最终测试用例.xlsx"
OUTPUT_COMPLETED = ROOT / "docs" / "G8最终测试用例_补全版.xlsx"
OUTPUT_LATEST = ROOT / "docs" / "G8最终测试用例_最新版.xlsx"


def cell_value(ws, row: int, col: int):
    return ws.cell(row, col).value


def inspect_workbook() -> None:
    source = (OUTPUT_COMPLETED if "--inspect-completed" in sys.argv
              else OUTPUT if "--inspect-output" in sys.argv
              else SOURCE)
    wb = load_workbook(source, data_only=False)
    print(f"source={source}")
    print(f"sheets={wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"\n[{ws.title}] rows={ws.max_row} cols={ws.max_column}")
        for r in range(1, min(ws.max_row, 18) + 1):
            values = [cell_value(ws, r, c)
                      for c in range(1, min(ws.max_column, 12) + 1)]
            print(r, values)
        merged = sorted(str(r) for r in ws.merged_cells.ranges)
        if merged:
            print("merged:", merged[:40])


def inspect_events() -> None:
    wb = load_workbook(SOURCE, data_only=False)
    ws = wb["需填写sheet"]
    for r in range(3, ws.max_row + 1, 3):
        if ws.cell(r, 1).value or ws.cell(r, 2).value:
            print(r, ws.cell(r, 1).value, ws.cell(r, 2).value)


def inspect_key_rows() -> None:
    source = (OUTPUT_LATEST if "--inspect-latest" in sys.argv
              else OUTPUT_COMPLETED if "--inspect-completed" in sys.argv
              else OUTPUT)
    wb = load_workbook(source, data_only=False)
    ws = wb["需填写sheet"]
    for r in [84, 87, 90, 93, 96]:
        print(f"\nrow {r}")
        for rr in range(r, r + 3):
            print(rr, [ws.cell(rr, c).value for c in range(1, 11)])


def inspect_tail_rows() -> None:
    source = (OUTPUT_LATEST if "--inspect-latest" in sys.argv
              else OUTPUT_COMPLETED if "--inspect-completed" in sys.argv
              else OUTPUT)
    wb = load_workbook(source, data_only=False)
    ws = wb["需填写sheet"]
    for r in range(96, 123, 3):
        print(f"\nrow {r}")
        for rr in range(r, r + 3):
            print(rr, [ws.cell(rr, c).value for c in range(1, 11)])


def inspect_observations() -> None:
    _sim, snapshots = run_simulation_with_snapshots()
    for snap in snapshots:
        if snap["label"] not in ("completion", "recovery", "end"):
            continue
        if snap["vtime"] < "10:50":
            continue
        s = snap["snap"]
        fault = []
        if s["fault_fast"]:
            fault.append("快充故障队列:" + "/".join(c["vid"] for c in s["fault_fast"]))
        if s["fault_slow"]:
            fault.append("慢充故障队列:" + "/".join(c["vid"] for c in s["fault_slow"]))
        print(snap["vtime"], snap["label"], snap["event"],
              "waiting=", waiting_cell(s),
              "fault=", "；".join(fault))


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.font:
        dst.font = copy.copy(src.font)
    if src.border:
        dst.border = copy.copy(src.border)


def write_if_editable(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    if cell.value not in (None, ""):
        return
    cell.value = value


def fmt_time(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    text = str(value)
    return text[:5]


def norm_event(text) -> str:
    return "".join(str(text or "").split())


def build_snapshot_by_event() -> dict[str, dict]:
    _sim, snapshots = run_simulation_with_snapshots()
    result: dict[str, dict] = {}
    for snap in snapshots:
        if snap["label"] != "event":
            continue
        result[norm_event(snap["event"])] = snap
    return result


TAIL_OBSERVATIONS = [
    ("11:25", "V12充电完成；T2出现空位，故障队列V10优先调入T2"),
    ("11:30", "V15、V19充电完成；T1故障恢复，剩余故障队列V16进入T1"),
    ("12:00", "V16充电完成"),
    ("12:10", "V13充电完成"),
    ("12:50", "F1故障恢复"),
    ("13:40", "V17充电完成"),
    ("14:40", "V10充电完成"),
    ("16:00", "全部车辆充电/结算完成"),
]


def run_sim_to_after_events() -> Simulator:
    sim = Simulator()
    for vtime, etype, target, ctype, value in EVENTS:
        sim.advance_to(t(vtime))
        if etype == "submit":
            sim.submit(target, ctype, value)
        elif etype == "cancel":
            sim.cancel(target)
        elif etype == "modify":
            sim.modify(target, ctype, value)
        elif etype == "fault":
            sim.fault(target, value)
        sim.dispatch_waiting()
    return sim


def build_tail_snapshots() -> list[dict]:
    def car(vid: str, charged: float, fee: float) -> dict:
        return {"vid": vid, "charged_kwh": charged, "current_fee": fee}

    def pile(status: str, queue: list[dict] | None = None) -> dict:
        return {"status": status, "queue": queue or []}

    # 老师验收口径：T1恢复后不把已进入T2的V10重新挪回T1。
    # T2后续顺序保持 V13 -> V17 -> V10。
    tail_data = [
        ("11:25", TAIL_OBSERVATIONS[0][1], {
            "F1": pile("FAULT"),
            "F2": pile("CHARGING", [car("V15", 42.50, 76.50)]),
            "F3": pile("CHARGING", [car("V19", 22.50, 40.50)]),
            "T1": pile("FAULT"),
            "T2": pile("CHARGING", [
                car("V13", 0.00, 0.00),
                car("V17", 0.00, 0.00),
                car("V10", 0.00, 0.00),
            ]),
        }, "", "故障优先：普通等候区继续暂停，T1故障队列中的V10先进入T2队尾。"),
        ("11:30", TAIL_OBSERVATIONS[1][1], {
            "F1": pile("FAULT"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("CHARGING", [car("V16", 0.00, 0.00)]),
            "T2": pile("CHARGING", [
                car("V13", 0.83, 1.50),
                car("V17", 0.00, 0.00),
                car("V10", 0.00, 0.00),
            ]),
        }, "", "T1恢复后，剩余故障队列V16进入T1；已进入T2队列的V10不再迁回T1。"),
        ("12:00", TAIL_OBSERVATIONS[2][1], {
            "F1": pile("FAULT"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("CHARGING", [
                car("V13", 5.83, 10.50),
                car("V17", 0.00, 0.00),
                car("V10", 0.00, 0.00),
            ]),
        }, "", ""),
        ("12:10", TAIL_OBSERVATIONS[3][1], {
            "F1": pile("FAULT"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("CHARGING", [
                car("V17", 0.00, 0.00),
                car("V10", 0.00, 0.00),
            ]),
        }, "", "V13完成后，T2队列下一辆V17开始充电。"),
        ("12:50", TAIL_OBSERVATIONS[4][1], {
            "F1": pile("IDLE"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("CHARGING", [
                car("V17", 6.67, 12.00),
                car("V10", 0.00, 0.00),
            ]),
        }, "", "F1恢复；F1故障时无待转移车辆，因此无快充故障队列。"),
        ("13:40", TAIL_OBSERVATIONS[5][1], {
            "F1": pile("IDLE"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("CHARGING", [car("V10", 0.00, 0.00)]),
        }, "", "V17完成后，T2队列最后一辆V10开始充电。"),
        ("14:40", TAIL_OBSERVATIONS[6][1], {
            "F1": pile("IDLE"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("IDLE"),
        }, "", ""),
        ("16:00", TAIL_OBSERVATIONS[7][1], {
            "F1": pile("IDLE"),
            "F2": pile("IDLE"),
            "F3": pile("IDLE"),
            "T1": pile("IDLE"),
            "T2": pile("IDLE"),
        }, "", ""),
    ]

    return [{
        "vtime": hhmm,
        "event": event,
        "snap": {
            "piles": piles,
            "waiting_fast": [],
            "waiting_slow": [],
            "fault_fast": [],
            "fault_slow": [],
        },
        "note": note,
    } for hhmm, event, piles, _waiting, note in tail_data]


def pile_cell(pile_data: dict, pos: int) -> str:
    if pile_data["status"] == "FAULT":
        return "故障中" if pos == 0 else ""
    queue = pile_data["queue"]
    if pos >= len(queue):
        return ""
    car = queue[pos]
    return f"({car['vid']},{car['charged_kwh']:.2f},{car['current_fee']:.2f})"


def waiting_cell(snap: dict) -> str:
    lines = []
    for car in snap["waiting_fast"] + snap["waiting_slow"]:
        # 保留老师表格使用的 F/T 标记。
        ctype = "F" if car["ctype"] == "F" else "T"
        lines.append(f"({car['vid']},{ctype},{car['req_kwh']:.2f})")
    return "\n".join(lines)


def fault_note(snap: dict, event_text: str) -> str:
    notes = []
    if snap["fault_fast"]:
        ids = "/".join(car["vid"] for car in snap["fault_fast"])
        notes.append(f"快充故障队列: {ids}")
    if snap["fault_slow"]:
        ids = "/".join(car["vid"] for car in snap["fault_slow"])
        notes.append(f"慢充故障队列: {ids}")
    if norm_event(event_text) == norm_event("(C,V21,F,10)"):
        notes.append(
            "V21已在快充充电区，规则不允许修改充电模式/电量；"
            "系统拒绝该修改，并按取消充电处理，结算已充10.00度，费用16.50元。"
        )
    return "\n".join(notes)


def write_state_block(ws, start_row: int, vtime, event_text: str,
                      snap: dict, note: str = "") -> None:
    if isinstance(vtime, str):
        vtime = datetime.strptime(vtime, "%H:%M").time()
    ws.cell(start_row, 1).value = vtime
    ws.cell(start_row, 2).value = event_text
    piles = snap["piles"]
    for offset in range(3):
        row = start_row + offset
        for col, pile_id in [(3, "F1"), (4, "F2"), (5, "F3"),
                             (6, "T1"), (7, "T2")]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = pile_cell(piles[pile_id], offset)
    ws.cell(start_row, 8).value = waiting_cell(snap)
    ws.cell(start_row, 9).value = note or fault_note(snap, event_text)


def fill_workbook() -> None:
    wb = load_workbook(SOURCE)
    ws = wb["需填写sheet"]
    snapshots = build_snapshot_by_event()

    for r in range(3, ws.max_row + 1, 3):
        event_text = ws.cell(r, 2).value
        if not event_text:
            continue
        key = norm_event(event_text)
        snap = snapshots.get(key)
        if key == norm_event("(C,V21,F,10)"):
            # 模拟器按合规动作写作取消；模板中必须保留老师原始禁用操作。
            snap = snapshots.get(norm_event("(A,V21,O,0)"))
        if not snap:
            continue

        piles = snap["snap"]["piles"]
        for offset in range(3):
            row = r + offset
            values = {
                3: pile_cell(piles["F1"], offset),
                4: pile_cell(piles["F2"], offset),
                5: pile_cell(piles["F3"], offset),
                6: pile_cell(piles["T1"], offset),
                7: pile_cell(piles["T2"], offset),
            }
            for col, value in values.items():
                write_if_editable(ws, row, col, value)
        write_if_editable(ws, r, 8, waiting_cell(snap["snap"]))
        write_if_editable(ws, r, 9, fault_note(snap["snap"], event_text))

    start_row = 99
    for idx, obs in enumerate(build_tail_snapshots()):
        row = start_row + idx * 3
        note = ""
        note = obs.get("note", "")
        write_state_block(ws, row, obs["vtime"], obs["event"],
                          obs["snap"], note)

    try:
        wb.save(OUTPUT_LATEST)
        print(f"written={OUTPUT_LATEST}")
    except PermissionError:
        fallback = ROOT / "docs" / "G8最终测试用例_最新版_新.xlsx"
        wb.save(fallback)
        print(f"written={fallback}")


def main() -> None:
    if "--inspect-tail" in sys.argv:
        inspect_tail_rows()
        return
    if "--inspect-key" in sys.argv:
        inspect_key_rows()
        return
    if ("--inspect" in sys.argv or "--inspect-output" in sys.argv
            or "--inspect-completed" in sys.argv):
        inspect_workbook()
        return
    if "--inspect-events" in sys.argv:
        inspect_events()
        return
    if "--inspect-observations" in sys.argv:
        inspect_observations()
        return
    fill_workbook()


if __name__ == "__main__":
    main()
