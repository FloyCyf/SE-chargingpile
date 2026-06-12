from __future__ import annotations

import copy
import sys
from pathlib import Path
from datetime import time

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "docs" / "G8测试用例.xlsx"
OUTPUT = ROOT / "docs" / "G9测试用例.xlsx"


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.border:
            dst.border = copy.copy(src.border)


def clear_data_rows(ws, first_data_row: int = 2) -> None:
    for row in range(first_data_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def write_rows_like_template(ws, rows: list[list[object]]) -> None:
    first_data_row = 2
    clear_data_rows(ws, first_data_row)
    template_row = first_data_row
    for idx, row_values in enumerate(rows, start=first_data_row):
        if idx > ws.max_row:
            ws.insert_rows(idx)
        copy_row_style(ws, template_row, idx)
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(idx, col)
            if not isinstance(cell, MergedCell):
                cell.value = value


def clear_body(ws, start_row: int = 3) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= start_row:
            ws.unmerge_cells(str(merged))
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def fmt_wait(items: list[tuple[str, str, float]]) -> str:
    if not items:
        return ""
    return "；".join(f"({v},{mode},{k:g})" for v, mode, k in items)


def fmt_pile_item(item) -> str:
    if item is None:
        return ""
    if isinstance(item, str):
        return item
    vehicle_id, charged, fee = item
    return f"({vehicle_id},{charged:.2f},{fee:.2f})"


def write_event_block(ws, start_row: int, event: dict) -> None:
    for offset in range(3):
        src_row = 3 + offset
        dst_row = start_row + offset
        copy_row_style(ws, src_row, dst_row)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(dst_row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row + 2, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2,
                   end_row=start_row + 2, end_column=2)
    ws.merge_cells(start_row=start_row, start_column=8,
                   end_row=start_row + 2, end_column=8)
    ws.merge_cells(start_row=start_row, start_column=9,
                   end_row=start_row + 2, end_column=10)

    ws.cell(start_row, 1).value = event["time"]
    ws.cell(start_row, 2).value = event["event"]
    ws.cell(start_row, 8).value = event.get("waiting", "")
    ws.cell(start_row, 9).value = event.get("note", "")

    for col_idx, key in [(3, "F1"), (4, "F2"), (5, "F3"),
                         (6, "T1"), (7, "T2")]:
        values = event.get(key, [])
        for offset in range(3):
            ws.cell(start_row + offset, col_idx).value = fmt_pile_item(
                values[offset] if offset < len(values) else None)


def main() -> None:
    if "--inspect-output" in sys.argv:
        out_wb = load_workbook(OUTPUT, data_only=True)
        out_ws = out_wb.worksheets[0]
        for r in range(1, min(out_ws.max_row, 35) + 1):
            print(r, [out_ws.cell(r, c).value for c in range(1, min(out_ws.max_column, 10) + 1)])
        print("merged:", sorted(str(r) for r in out_ws.merged_cells.ranges)[:80])
        return

    template_path = (ROOT / "docs" / "G8测试用例_修正版.xlsx"
                     if "--inspect-revised-events" in sys.argv
                     else TEMPLATE)
    wb = load_workbook(template_path)

    # G8 工作簿第一张表是验收事件表；保留原格式，只替换为 G9 用例。
    ws = wb.worksheets[0]
    if "--inspect-template" in sys.argv:
        for r in range(1, min(ws.max_row, 20) + 1):
            print(r, [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)])
        merged = sorted(str(r) for r in ws.merged_cells.ranges)
        print("merged:", merged[:80])
        return

    if ("--inspect-template-events" in sys.argv
            or "--inspect-revised-events" in sys.argv):
        for r in range(3, ws.max_row + 1, 3):
            print(r, ws.cell(r, 1).value, ws.cell(r, 2).value,
                  "wait=", ws.cell(r, 8).value,
                  "fault=", ws.cell(r, 9).value)
        return

    ws.title = "G9测试用例"

    phase1 = [
        ("V1", "Fast", 3.0), ("V2", "Fast", 15.0), ("V3", "Fast", 20.0),
        ("V4", "Fast", 3.0), ("V5", "Fast", 12.0), ("V6", "Fast", 18.0),
        ("V7", "Fast", 3.0), ("V8", "Fast", 10.0), ("V9", "Fast", 15.0),
        ("V10", "Slow", 3.0), ("V11", "Slow", 6.0), ("V12", "Slow", 6.0),
        ("V13", "Slow", 3.0), ("V14", "Slow", 6.0), ("V15", "Slow", 6.0),
    ]
    phase2 = [
        ("V16", "Fast", 25.0), ("V17", "Fast", 10.0),
        ("V18", "Fast", 5.0), ("V19", "Fast", 12.0),
        ("V20", "Fast", 30.0),
        ("V21", "Slow", 25.0), ("V22", "Slow", 18.0),
        ("V23", "Slow", 4.0), ("V24", "Slow", 22.0),
    ]

    events = [
        {
            "time": time(6, 0),
            "event": "初始化：3快2慢；每桩3车位；等候区10；策略=batch_min_total",
            "note": "与 scripts/g9_test.py 保持一致，虚拟时钟 06:00，ratio=4。",
        },
        {
            "time": time(6, 0),
            "event": "(A,V1,F,3)(A,V4,F,3)(A,V7,F,3)(A,V10,T,3)(A,V13,T,3)",
            "F1": [("V1", 0, 0)],
            "F2": [("V4", 0, 0)],
            "F3": [("V7", 0, 0)],
            "T1": [("V10", 0, 0)],
            "T2": [("V13", 0, 0)],
            "note": "先提交每个桩 position 0 的 3kWh 小车，保证后续能同时释放多个空位。",
        },
        {
            "time": time(6, 0),
            "event": "(A,V2,F,15)(A,V5,F,12)(A,V8,F,10)(A,V11,T,6)(A,V14,T,6)",
            "F1": [("V1", 0, 0), ("V2", 0, 0)],
            "F2": [("V4", 0, 0), ("V5", 0, 0)],
            "F3": [("V7", 0, 0), ("V8", 0, 0)],
            "T1": [("V10", 0, 0), ("V11", 0, 0)],
            "T2": [("V13", 0, 0), ("V14", 0, 0)],
            "note": "继续填充各桩 position 1。",
        },
        {
            "time": time(6, 0),
            "event": "(A,V3,F,20)(A,V6,F,18)(A,V9,F,15)(A,V12,T,6)(A,V15,T,6)",
            "F1": [("V1", 0, 0), ("V2", 0, 0), ("V3", 0, 0)],
            "F2": [("V4", 0, 0), ("V5", 0, 0), ("V6", 0, 0)],
            "F3": [("V7", 0, 0), ("V8", 0, 0), ("V9", 0, 0)],
            "T1": [("V10", 0, 0), ("V11", 0, 0), ("V12", 0, 0)],
            "T2": [("V13", 0, 0), ("V14", 0, 0), ("V15", 0, 0)],
            "note": "充电区 5 个桩均满。",
        },
        {
            "time": time(6, 0),
            "event": "(A,V16,F,25)(A,V17,F,10)(A,V18,F,5)(A,V19,F,12)(A,V20,F,30)(A,V21,T,25)(A,V22,T,18)(A,V23,T,4)(A,V24,T,22)",
            "F1": [("V1", 0, 0), ("V2", 0, 0), ("V3", 0, 0)],
            "F2": [("V4", 0, 0), ("V5", 0, 0), ("V6", 0, 0)],
            "F3": [("V7", 0, 0), ("V8", 0, 0), ("V9", 0, 0)],
            "T1": [("V10", 0, 0), ("V11", 0, 0), ("V12", 0, 0)],
            "T2": [("V13", 0, 0), ("V14", 0, 0), ("V15", 0, 0)],
            "waiting": fmt_wait(phase2),
            "note": "9 辆进入等候区，等待批量调度。",
        },
        {
            "time": time(6, 6),
            "event": "快充 position0 完成；触发第一次 batch 调度",
            "F1": [("V2", 0, 0), ("V3", 0, 0), "V18(5kWh)"],
            "F2": [("V5", 0, 0), ("V6", 0, 0), "V17(10kWh)"],
            "F3": [("V8", 0, 0), ("V9", 0, 0), "V19(12kWh)"],
            "T1": [("V10", 1.0, 1.2), ("V11", 0, 0), ("V12", 0, 0)],
            "T2": [("V13", 1.0, 1.2), ("V14", 0, 0), ("V15", 0, 0)],
            "waiting": fmt_wait([
                ("V16", "Fast", 25.0), ("V20", "Fast", 30.0),
                ("V21", "Slow", 25.0), ("V22", "Slow", 18.0),
                ("V23", "Slow", 4.0), ("V24", "Slow", 22.0),
            ]),
            "note": "优越性观察点：batch 叫 V18/V17/V19；FIFO 会叫 V16/V17/V18。",
        },
        {
            "time": time(6, 18),
            "event": "慢充 position0 完成；触发慢充 batch 调度",
            "F1": [("V2", 6.0, 7.2), ("V3", 0, 0), "V18(5kWh)"],
            "F2": [("V5", 6.0, 7.2), ("V6", 0, 0), "V17(10kWh)"],
            "F3": [("V8", 6.0, 7.2), ("V9", 0, 0), "V19(12kWh)"],
            "T1": [("V11", 0, 0), ("V12", 0, 0), "V23(4kWh)"],
            "T2": [("V14", 0, 0), ("V15", 0, 0), "V22(18kWh)"],
            "waiting": fmt_wait([
                ("V16", "Fast", 25.0), ("V20", "Fast", 30.0),
                ("V21", "Slow", 25.0), ("V24", "Slow", 22.0),
            ]),
            "note": "慢充按总完成时长最短，优先 V23(4) 与 V22(18)。",
        },
        {
            "time": time(6, 26),
            "event": "快充后续空位释放；调度剩余快充等待车",
            "F1": [("V2", 10.0, 12.0), ("V3", 0, 0), "V18(5kWh)"],
            "F2": [("V5", 10.0, 12.0), ("V6", 0, 0), "V17(10kWh)"],
            "F3": [("V9", 0, 0), "V19(12kWh)", "V16(25kWh)"],
            "T1": [("V11", 1.33, 1.6), ("V12", 0, 0), "V23(4kWh)"],
            "T2": [("V14", 1.33, 1.6), ("V15", 0, 0), "V22(18kWh)"],
            "waiting": fmt_wait([
                ("V20", "Fast", 30.0),
                ("V21", "Slow", 25.0), ("V24", "Slow", 22.0),
            ]),
            "note": "V16 进入快充桩；后续 F2/F1 释放时 V20 进入。",
        },
        {
            "time": time(6, 54),
            "event": "慢充再次释放空位；调度 V24、V21",
            "F1": ["快充队列继续推进"],
            "F2": ["快充队列继续推进"],
            "F3": ["快充队列继续推进"],
            "T1": [("V12", 0, 0), "V23(4kWh)", "V24(22kWh)"],
            "T2": [("V15", 0, 0), "V22(18kWh)", "V21(25kWh)"],
            "waiting": "",
            "note": "V24(22) 比 V21(25) 更短，先进入调度集合；等候区清空。",
        },
        {
            "time": time(7, 0),
            "event": "结果核对",
            "F1": ["无慢充车进入快充桩"],
            "F2": ["无慢充车进入快充桩"],
            "F3": ["无慢充车进入快充桩"],
            "T1": ["无快充车进入慢充桩"],
            "T2": ["无快充车进入慢充桩"],
            "waiting": "",
            "note": "核心判定：按充电模式分配；多个空位时一次叫多号；非 FIFO，短完成时长车辆优先。",
        },
    ]

    clear_body(ws)
    for idx, event in enumerate(events):
        write_event_block(ws, 3 + idx * 3, event)

    last_row = 2 + len(events) * 3
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

    # 如果模板还有其它说明 sheet，保留格式并更新标题中明显的 G8 字样。
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("G8", "G9")

    wb.save(OUTPUT)
    check_wb = load_workbook(OUTPUT, data_only=True)
    check_ws = check_wb.worksheets[0]
    print(OUTPUT)
    print(f"sheet={check_ws.title}, rows={check_ws.max_row}, cols={check_ws.max_column}")
    print(f"A1:J1={ [check_ws.cell(1, c).value for c in range(1, min(10, check_ws.max_column) + 1)] }")
    print(f"A2:J2={ [check_ws.cell(2, c).value for c in range(1, min(10, check_ws.max_column) + 1)] }")


if __name__ == "__main__":
    main()
