from datetime import datetime, timedelta
import csv
import io
import openpyxl
from fastapi import APIRouter, Request, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from src.api.auth import require_admin
from src.api.schemas import (
    PileControlRequest, PileControlResponse,
    SystemStatusResponse, WaitingAreaResponse,
    ReportResponse, ReportItem,
    PileStatusLogItem, PileStatusLogResponse,
    BillItem, BillDetailItem, BillListResponse,
    AdminOrderItem, AdminOrderListResponse,
)
from src.models.database import AsyncSessionLocal
from src.models.models import ChargeOrder, OrderStatus, PileStatusLog, Bill, BillDetail
from src.core.billing import get_billing_config, update_billing_config

router = APIRouter()


@router.post("/piles/{pile_id}/control", response_model=PileControlResponse)
async def control_pile(
    pile_id: str,
    body: PileControlRequest,
    request: Request,
    admin: dict = Depends(require_admin),
):
    """启动/关闭/故障 充电桩"""
    scheduler = request.app.state.scheduler
    action = body.action.lower()

    if action == "start":
        result = await scheduler.start_pile(pile_id)
    elif action == "stop":
        result = await scheduler.stop_pile(pile_id)
    elif action == "fault":
        result = await scheduler.fault_pile(
            pile_id, duration_minutes=body.duration_minutes)
    elif action == "recover":
        result = await scheduler.recover_pile(pile_id)
    else:
        raise HTTPException(status_code=400,
                            detail="无效操作，可选: start/stop/fault/recover")

    if result['status'] == 'failed':
        raise HTTPException(status_code=400, detail=result['message'])
    return PileControlResponse(**result)


@router.get("/piles", response_model=SystemStatusResponse)
async def get_all_piles(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """获取所有充电桩状态（含队列详情）"""
    scheduler = request.app.state.scheduler
    status = scheduler.get_system_status()
    return SystemStatusResponse(**status)


@router.get("/waiting-area", response_model=WaitingAreaResponse)
async def get_waiting_area(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """获取等候区车辆信息"""
    scheduler = request.app.state.scheduler
    data = scheduler.get_waiting_area()
    return WaitingAreaResponse(**data)


@router.post("/dispatch")
async def manual_dispatch(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """手动触发等候区 → 桩队列调度"""
    scheduler = request.app.state.scheduler
    async with scheduler.lock:
        await scheduler.dispatch_from_waiting_area_async()
    return {"status": "success", "message": "调度完成"}


@router.get("/reports", response_model=ReportResponse)
async def get_reports(
    request: Request,
    period: str = Query("day", description="统计周期: day/week/month"),
    date: str = Query(None, description="日期 YYYY-MM-DD，默认今天"),
    admin: dict = Depends(require_admin),
):
    """生成充电桩报表"""
    # 解析日期范围
    if date:
        try:
            base_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，应为 YYYY-MM-DD")
    else:
        base_date = datetime.now()

    if period == "day":
        start_dt = base_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = start_dt + timedelta(days=1)
    elif period == "week":
        # 本周一 ~ 下周一
        weekday = base_date.weekday()
        start_dt = (base_date - timedelta(days=weekday)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        end_dt = start_dt + timedelta(days=7)
    elif period == "month":
        start_dt = base_date.replace(
            day=1, hour=0, minute=0, second=0, microsecond=0)
        if start_dt.month == 12:
            end_dt = start_dt.replace(year=start_dt.year + 1, month=1)
        else:
            end_dt = start_dt.replace(month=start_dt.month + 1)
    else:
        raise HTTPException(status_code=400,
                            detail="无效周期，可选: day/week/month")

    # 从数据库聚合已完成订单
    async with AsyncSessionLocal() as session:
        stmt = (
            select(
                ChargeOrder.pile_id,
                ChargeOrder.charge_type,
                func.count(ChargeOrder.id).label("charge_count"),
                func.coalesce(func.sum(ChargeOrder.charge_duration), 0.0).label("total_duration"),
                func.coalesce(func.sum(ChargeOrder.total_power), 0.0).label("total_kwh"),
                func.coalesce(func.sum(ChargeOrder.power_fee), 0.0).label("total_power_fee"),
                func.coalesce(func.sum(ChargeOrder.service_fee), 0.0).label("total_service_fee"),
                func.coalesce(func.sum(ChargeOrder.total_fee), 0.0).label("total_total_fee"),
            )
            .where(and_(
                ChargeOrder.status.in_([
                    OrderStatus.COMPLETED, OrderStatus.FAULTED]),
                ChargeOrder.finished_at >= start_dt,
                ChargeOrder.finished_at < end_dt,
                ChargeOrder.pile_id.isnot(None),
            ))
            .group_by(ChargeOrder.pile_id, ChargeOrder.charge_type)
            .order_by(ChargeOrder.pile_id)
        )
        result = await session.execute(stmt)
        rows = result.all()

    items = []
    for row in rows:
        items.append(ReportItem(
            pile_id=row.pile_id or "",
            pile_type=row.charge_type or "",
            charge_count=row.charge_count,
            total_duration=round(row.total_duration, 4),
            total_kwh=round(row.total_kwh, 4),
            total_power_fee=round(row.total_power_fee, 2),
            total_service_fee=round(row.total_service_fee, 2),
            total_total_fee=round(row.total_total_fee, 2),
        ))

    return ReportResponse(
        period=period,
        start_date=start_dt.strftime("%Y-%m-%d"),
        end_date=end_dt.strftime("%Y-%m-%d"),
        items=items,
    )


@router.get("/pile-status-logs", response_model=PileStatusLogResponse)
async def get_pile_status_logs(
    request: Request,
    pile_id: str = Query(None, description="充电桩编号，为空则查询全部"),
    limit: int = Query(50, ge=1, le=500, description="返回条数"),
    offset: int = Query(0, ge=0, description="偏移量"),
    admin: dict = Depends(require_admin),
):
    """查询充电桩状态变更历史"""
    async with AsyncSessionLocal() as session:
        stmt = select(PileStatusLog).order_by(
            PileStatusLog.changed_at.desc())

        if pile_id:
            stmt = stmt.where(PileStatusLog.pile_id == pile_id)

        # 查总数
        count_stmt = select(func.count()).select_from(PileStatusLog)
        if pile_id:
            count_stmt = count_stmt.where(PileStatusLog.pile_id == pile_id)
        total_result = await session.execute(count_stmt)
        total = total_result.scalar() or 0

        stmt = stmt.offset(offset).limit(limit)
        result = await session.execute(stmt)
        logs = result.scalars().all()

    return PileStatusLogResponse(
        logs=[
            PileStatusLogItem(
                id=log.id,
                pile_id=log.pile_id,
                old_status=log.old_status,
                new_status=log.new_status,
                reason=log.reason,
                operator=log.operator or "system",
                changed_at=log.changed_at,
            ) for log in logs
        ],
        total=total,
    )


# ---- 计费配置 ----

class BillingConfigUpdate(BaseModel):
    peak_rate: Optional[float] = None
    flat_rate: Optional[float] = None
    valley_rate: Optional[float] = None
    service_fee_rate: Optional[float] = None
    peak_hours: Optional[List[List[int]]] = None
    flat_hours: Optional[List[List[int]]] = None
    valley_hours: Optional[List[List[int]]] = None


@router.get("/billing-config")
async def get_billing_config_api(
    admin: dict = Depends(require_admin),
):
    """获取当前计费配置"""
    return get_billing_config()


@router.put("/billing-config")
async def update_billing_config_api(
    body: BillingConfigUpdate,
    admin: dict = Depends(require_admin),
):
    """动态更新计费配置（费率/时段）"""
    new_config = body.model_dump(exclude_none=True)
    if not new_config:
        raise HTTPException(status_code=400, detail="未提供任何更新字段")
    update_billing_config(new_config)
    return {"status": "success", "message": "计费配置已更新", "config": get_billing_config()}


# ---- 订单列表 ----

@router.get("/orders", response_model=AdminOrderListResponse)
async def admin_list_orders(
    status: str = Query(None, description="筛选状态"),
    vehicle_id: str = Query(None, description="筛选车牌号"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    admin: dict = Depends(require_admin),
):
    """管理员查询所有订单"""
    async with AsyncSessionLocal() as session:
        stmt = select(ChargeOrder).order_by(ChargeOrder.id.desc())
        count_stmt = select(func.count()).select_from(ChargeOrder)

        if status:
            stmt = stmt.where(ChargeOrder.status == status)
            count_stmt = count_stmt.where(ChargeOrder.status == status)
        if vehicle_id:
            stmt = stmt.where(ChargeOrder.vehicle_id == vehicle_id)
            count_stmt = count_stmt.where(ChargeOrder.vehicle_id == vehicle_id)

        total = (await session.execute(count_stmt)).scalar() or 0
        result = await session.execute(stmt.offset(offset).limit(limit))
        orders = result.scalars().all()

    return AdminOrderListResponse(
        orders=[
            AdminOrderItem(
                order_id=o.id,
                vehicle_id=o.vehicle_id,
                charge_type=o.charge_type,
                requested_kwh=o.requested_kwh or 0.0,
                charged_kwh=o.charged_kwh or 0.0,
                queue_number=o.queue_number,
                status=o.status,
                pile_id=o.pile_id,
                total_fee=o.total_fee,
                created_at=o.created_at,
                started_at=o.started_at,
                finished_at=o.finished_at,
            ) for o in orders
        ],
        total=total,
    )


# ---- 账单列表（含详单） ----

@router.get("/bills", response_model=BillListResponse)
async def admin_list_bills(
    vehicle_id: str = Query(None, description="筛选车牌号"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    admin: dict = Depends(require_admin),
):
    """管理员查询所有账单（含详单明细）"""
    async with AsyncSessionLocal() as session:
        stmt = (select(Bill)
                .options(selectinload(Bill.details))
                .order_by(Bill.id.desc()))
        count_stmt = select(func.count()).select_from(Bill)

        if vehicle_id:
            stmt = stmt.where(Bill.vehicle_id == vehicle_id)
            count_stmt = count_stmt.where(Bill.vehicle_id == vehicle_id)

        total = (await session.execute(count_stmt)).scalar() or 0
        result = await session.execute(stmt.offset(offset).limit(limit))
        bills = result.scalars().unique().all()

    return BillListResponse(
        bills=[
            BillItem(
                id=b.id,
                bill_code=b.bill_code,
                order_id=b.order_id,
                vehicle_id=b.vehicle_id,
                pile_id=b.pile_id,
                charge_type=b.charge_type,
                charge_start_time=b.charge_start_time,
                charge_end_time=b.charge_end_time,
                charge_duration=b.charge_duration,
                total_power=b.total_power,
                power_fee=b.power_fee,
                service_fee=b.service_fee,
                total_fee=b.total_fee,
                created_at=b.created_at,
                details=[
                    BillDetailItem(
                        id=d.id,
                        period=d.period,
                        start_time=d.start_time or "",
                        end_time=d.end_time or "",
                        duration_minutes=d.duration_minutes or 0,
                        kwh=d.kwh or 0.0,
                        rate=d.rate or 0.0,
                        fee=d.fee or 0.0,
                    ) for d in (b.details or [])
                ],
            ) for b in bills
        ],
        total=total,
    )


# ---- CSV 导出 ----

@router.get("/export/orders")
async def export_orders_csv(
    admin: dict = Depends(require_admin),
):
    """导出所有订单为 CSV 文件"""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(ChargeOrder).order_by(ChargeOrder.id.desc()))
        orders = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "订单ID", "车牌号", "充电桩", "充电模式", "请求电量(kWh)",
        "实际电量(kWh)", "排队号", "状态", "充电费(元)", "服务费(元)",
        "总费用(元)", "创建时间", "开始充电", "完成时间",
    ])
    for o in orders:
        writer.writerow([
            o.id, o.vehicle_id, o.pile_id or "", o.charge_type,
            o.requested_kwh or 0, o.charged_kwh or 0,
            o.queue_number or "", o.status,
            o.power_fee or 0, o.service_fee or 0, o.total_fee or 0,
            o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else "",
            o.charge_start_time.strftime("%Y-%m-%d %H:%M:%S") if o.charge_start_time else "",
            o.finished_at.strftime("%Y-%m-%d %H:%M:%S") if o.finished_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=orders.csv"},
    )


@router.get("/export/bills")
async def export_bills_csv(
    admin: dict = Depends(require_admin),
):
    """导出所有账单为 CSV 文件"""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Bill)
            .options(selectinload(Bill.details))
            .order_by(Bill.id.desc()))
        bills = result.scalars().unique().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "账单编号", "订单ID", "车牌号", "充电桩", "充电模式",
        "充电时长(h)", "充电电量(kWh)", "充电费(元)", "服务费(元)",
        "总费用(元)", "开始时间", "结束时间", "账单生成时间",
    ])
    for b in bills:
        writer.writerow([
            b.bill_code, b.order_id, b.vehicle_id, b.pile_id or "",
            b.charge_type, round(b.charge_duration or 0, 4),
            round(b.total_power or 0, 4),
            round(b.power_fee or 0, 2), round(b.service_fee or 0, 2),
            round(b.total_fee or 0, 2),
            b.charge_start_time.strftime("%Y-%m-%d %H:%M:%S") if b.charge_start_time else "",
            b.charge_end_time.strftime("%Y-%m-%d %H:%M:%S") if b.charge_end_time else "",
            b.created_at.strftime("%Y-%m-%d %H:%M:%S") if b.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=bills.csv"},
    )


@router.get("/export/bill-details")
async def export_bill_details_csv(
    admin: dict = Depends(require_admin),
):
    """导出所有详单为 CSV 文件"""
    async with AsyncSessionLocal() as session:
        stmt = (select(BillDetail, Bill.bill_code, Bill.vehicle_id)
                .join(Bill, BillDetail.bill_id == Bill.id)
                .order_by(BillDetail.id.desc()))
        result = await session.execute(stmt)
        rows = result.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "详单ID", "账单编号", "车牌号", "时段类型",
        "开始时刻", "结束时刻", "持续(分钟)",
        "电量(kWh)", "电价(元/kWh)", "费用(元)",
    ])
    for detail, bill_code, vehicle_id in rows:
        period_name = {"peak": "峰时", "flat": "平时", "valley": "谷时"}.get(
            detail.period, detail.period)
        writer.writerow([
            detail.id, bill_code, vehicle_id, period_name,
            detail.start_time or "", detail.end_time or "",
            detail.duration_minutes or 0,
            round(detail.kwh or 0, 4), round(detail.rate or 0, 2),
            round(detail.fee or 0, 2),
        ])

    output.seek(0)
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=bill_details.csv"},
    )


# ---- 虚拟时钟管理 ----

class ClockSetBody(BaseModel):
    datetime: Optional[str] = None   # "YYYY-MM-DD HH:MM:SS" 或 "HH:MM:SS"（今日日期）
    ratio: Optional[float] = None    # 时间推进倍率
    real_minutes: Optional[float] = None
    virtual_minutes: Optional[float] = None
    allowed_start: Optional[str] = None
    allowed_end: Optional[str] = None


@router.get("/clock")
async def get_clock(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """获取当前虚拟时间和倍率"""
    clock = request.app.state.scheduler.clock
    return {
        "current_virtual_time": clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "ratio": clock.ratio,
        "real_minutes": getattr(request.app.state, "clock_real_minutes", None),
        "virtual_minutes": getattr(request.app.state, "clock_virtual_minutes", None),
        "allowed_start": getattr(request.app.state, "clock_allowed_start", None),
        "allowed_end": getattr(request.app.state, "clock_allowed_end", None),
        "running": clock.running,
        "description": f"每真实1秒 = 虚拟{clock.ratio}分钟",
    }


@router.put("/clock")
async def set_clock(
    body: ClockSetBody,
    request: Request,
    admin: dict = Depends(require_admin),
):
    """设置虚拟时间或倍率（可单独设置其中一项）"""
    clock = request.app.state.scheduler.clock
    changed = []

    def _parse_admin_dt(raw: str):
        dt_str = raw.strip().replace("T", " ")
        if len(dt_str) <= 8:
            from datetime import date as _date
            today = _date.today().strftime("%Y-%m-%d")
            dt_str = f"{today} {dt_str}"
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue
        raise HTTPException(
            status_code=400,
            detail="时间格式错误，支持 HH:MM、HH:MM:SS 或 YYYY-MM-DD HH:MM")

    parsed_allowed_start = None
    parsed_allowed_end = None

    if body.allowed_start is not None:
        parsed_allowed_start = _parse_admin_dt(body.allowed_start)
        request.app.state.clock_allowed_start = parsed_allowed_start.strftime("%Y-%m-%d %H:%M:%S")
        changed.append(f"允许开始时间设为 {request.app.state.clock_allowed_start}")

    if body.allowed_end is not None:
        parsed_allowed_end = _parse_admin_dt(body.allowed_end)
        request.app.state.clock_allowed_end = parsed_allowed_end.strftime("%Y-%m-%d %H:%M:%S")
        changed.append(f"允许结束时间设为 {request.app.state.clock_allowed_end}")

    if body.datetime is not None:
        new_dt = _parse_admin_dt(body.datetime)
        clock.set_time(new_dt)
        changed.append(f"虚拟时间设为 {new_dt.strftime('%Y-%m-%d %H:%M:%S')}")
    elif parsed_allowed_start is not None:
        clock.set_time(parsed_allowed_start)
        changed.append(f"虚拟时间自动同步为允许开始时间 {parsed_allowed_start.strftime('%Y-%m-%d %H:%M:%S')}")

    if body.real_minutes is not None or body.virtual_minutes is not None:
        if body.real_minutes is None or body.virtual_minutes is None:
            raise HTTPException(status_code=400, detail="real_minutes 和 virtual_minutes 需要同时提供")
        if body.real_minutes <= 0 or body.virtual_minutes <= 0:
            raise HTTPException(status_code=400, detail="real_minutes 和 virtual_minutes 必须大于0")
        body.ratio = body.virtual_minutes / (body.real_minutes * 60.0)
        request.app.state.clock_real_minutes = body.real_minutes
        request.app.state.clock_virtual_minutes = body.virtual_minutes
        changed.append(f"比例尺设为 真实{body.real_minutes}分钟 = 虚拟{body.virtual_minutes}分钟")

    if body.ratio is not None:
        if body.ratio <= 0:
            raise HTTPException(status_code=400, detail="倍率必须大于0")
        clock.set_ratio(body.ratio)
        changed.append(f"倍率设为 {body.ratio}（每真实1秒=虚拟{body.ratio}分钟）")

    if not changed:
        raise HTTPException(status_code=400, detail="请提供 datetime 或 ratio 字段")

    return {
        "status": "success",
        "changed": changed,
        "current_virtual_time": clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "ratio": clock.ratio,
        "real_minutes": getattr(request.app.state, "clock_real_minutes", None),
        "virtual_minutes": getattr(request.app.state, "clock_virtual_minutes", None),
        "allowed_start": getattr(request.app.state, "clock_allowed_start", None),
        "allowed_end": getattr(request.app.state, "clock_allowed_end", None),
        "running": clock.running,
    }


@router.post("/clock/start")
async def start_clock(
    request: Request,
    admin: dict = Depends(require_admin),
):
    clock = request.app.state.scheduler.clock
    clock.start()
    return {
        "status": "success",
        "message": "系统虚拟时间已开始流逝",
        "current_virtual_time": clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "ratio": clock.ratio,
        "running": clock.running,
    }


@router.post("/clock/pause")
async def pause_clock(
    request: Request,
    admin: dict = Depends(require_admin),
):
    clock = request.app.state.scheduler.clock
    clock.pause()
    return {
        "status": "success",
        "message": "系统虚拟时间已暂停",
        "current_virtual_time": clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "ratio": clock.ratio,
        "running": clock.running,
    }


@router.post("/clock/reset")
async def reset_clock(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """重置虚拟时间为当前真实时间（倍率不变）"""
    clock = request.app.state.scheduler.clock
    clock.reset()
    return {
        "status": "success",
        "message": "虚拟时间已重置为当前真实时间",
        "current_virtual_time": clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "ratio": clock.ratio,
        "running": clock.running,
    }


# ---- 验收快照 ----

@router.get("/acceptance/snapshot")
async def acceptance_snapshot(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """
    验收快照：按验收 Excel 结构返回当前全站状态。
    返回每个桩的队列车辆（含当前费用）和等候区车辆。
    """
    scheduler = request.app.state.scheduler
    status = scheduler.get_system_status()
    waiting = scheduler.get_waiting_area()

    fast_piles = {}
    slow_piles = {}
    for p in status["piles"]:
        pile_info = {
            "pile_id": p["pile_id"],
            "type": p["type"],
            "status": p["status"],
            "power": p["power"],
            "queue": [
                {
                    "position": qi["position"],
                    "vehicle_id": qi["vehicle_id"],
                    "queue_number": qi["queue_number"],
                    "requested_kwh": qi["requested_kwh"],
                    "charged_kwh": qi["charged_kwh"],
                    "current_fee": qi.get("current_fee", 0.0),
                    "current_power_fee": qi.get("current_power_fee", 0.0),
                    "current_service_fee": qi.get("current_service_fee", 0.0),
                    "charge_start_time": qi.get("charge_start_time"),
                }
                for qi in p["queue_items"]
            ],
        }
        if p["type"] == "Fast":
            fast_piles[p["pile_id"]] = pile_info
        else:
            slow_piles[p["pile_id"]] = pile_info

    return {
        "current_virtual_time": scheduler.clock.get_time().strftime("%Y-%m-%d %H:%M:%S"),
        "fast_piles": fast_piles,
        "slow_piles": slow_piles,
        "waiting_area": {
            "fast_waiting": waiting["fast_waiting"],
            "slow_waiting": waiting["slow_waiting"],
            "fast_count": len(waiting["fast_waiting"]),
            "slow_count": len(waiting["slow_waiting"]),
        },
    }


# ---- 系统参数设置 ----

class SystemParamsUpdate(BaseModel):
    waiting_area_size: Optional[int] = None       # 等候区最大容量 N
    pile_queue_length: Optional[int] = None       # 每桩总车位数 M
    alpha: Optional[float] = None                 # 低电量权重
    beta: Optional[float] = None                  # 充电模式权重
    gamma: Optional[float] = None                 # 等待时间权重
    fast_type_weight: Optional[float] = None      # 快充 W_type
    slow_type_weight: Optional[float] = None      # 慢充 W_type


@router.get("/system-params")
async def get_system_params(
    request: Request,
    admin: dict = Depends(require_admin),
):
    """查询当前系统调度参数"""
    s = request.app.state.scheduler
    return {
        "waiting_area_size": s.waiting_capacity,
        "pile_queue_length": s.pile_queue_length,
        "alpha": s.priority_alpha,
        "beta": s.priority_beta,
        "gamma": s.priority_gamma,
        "fast_type_weight": s.priority_fast_weight,
        "slow_type_weight": s.priority_slow_weight,
        "fast_power": s.fast_power,
        "slow_power": s.slow_power,
    }


@router.put("/system-params")
async def update_system_params(
    body: SystemParamsUpdate,
    request: Request,
    admin: dict = Depends(require_admin),
):
    """运行时更新系统调度参数（等候区容量、桩队列长度、优先级权重）"""
    params = body.model_dump(exclude_none=True)
    if not params:
        raise HTTPException(status_code=400, detail="未提供任何参数")
    result = await request.app.state.scheduler.update_system_params(params)
    if result["status"] == "failed":
        raise HTTPException(status_code=400, detail=result["message"])
    return result


# ---- Excel 导出 ----

def _make_xlsx_response(wb: "openpyxl.Workbook", filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/orders.xlsx")
async def export_orders_xlsx(
    admin: dict = Depends(require_admin),
):
    """导出所有订单为 Excel 文件"""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(ChargeOrder).order_by(ChargeOrder.id.desc()))
        orders = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"
    headers = [
        "订单ID", "车牌号", "充电桩", "充电模式", "请求电量(kWh)",
        "实际电量(kWh)", "排队号", "状态", "充电费(元)", "服务费(元)",
        "总费用(元)", "创建时间", "开始充电", "完成时间",
    ]
    ws.append(headers)
    for o in orders:
        ws.append([
            o.id, o.vehicle_id, o.pile_id or "", o.charge_type,
            o.requested_kwh or 0, o.charged_kwh or 0,
            o.queue_number or "", o.status,
            o.power_fee or 0, o.service_fee or 0, o.total_fee or 0,
            o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else "",
            o.charge_start_time.strftime("%Y-%m-%d %H:%M:%S") if o.charge_start_time else "",
            o.finished_at.strftime("%Y-%m-%d %H:%M:%S") if o.finished_at else "",
        ])
    return _make_xlsx_response(wb, "orders.xlsx")


@router.get("/export/bills.xlsx")
async def export_bills_xlsx(
    admin: dict = Depends(require_admin),
):
    """导出所有账单为 Excel 文件"""
    from sqlalchemy.orm import selectinload as _sel
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Bill).options(_sel(Bill.details)).order_by(Bill.id.desc()))
        bills = result.scalars().unique().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账单列表"
    ws.append([
        "账单编号", "订单ID", "车牌号", "充电桩", "充电模式",
        "充电时长(h)", "充电电量(kWh)", "充电费(元)", "服务费(元)",
        "总费用(元)", "开始时间", "结束时间", "账单生成时间",
    ])
    for b in bills:
        ws.append([
            b.bill_code, b.order_id, b.vehicle_id, b.pile_id or "",
            b.charge_type, round(b.charge_duration or 0, 4),
            round(b.total_power or 0, 4),
            round(b.power_fee or 0, 2), round(b.service_fee or 0, 2),
            round(b.total_fee or 0, 2),
            b.charge_start_time.strftime("%Y-%m-%d %H:%M:%S") if b.charge_start_time else "",
            b.charge_end_time.strftime("%Y-%m-%d %H:%M:%S") if b.charge_end_time else "",
            b.created_at.strftime("%Y-%m-%d %H:%M:%S") if b.created_at else "",
        ])
    return _make_xlsx_response(wb, "bills.xlsx")


@router.get("/export/bill-details.xlsx")
async def export_bill_details_xlsx(
    admin: dict = Depends(require_admin),
):
    """导出所有详单为 Excel 文件（含峰平谷分段）"""
    async with AsyncSessionLocal() as session:
        stmt = (select(BillDetail, Bill.bill_code, Bill.vehicle_id)
                .join(Bill, BillDetail.bill_id == Bill.id)
                .order_by(BillDetail.id.desc()))
        result = await session.execute(stmt)
        rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "详单列表"
    ws.append([
        "详单ID", "账单编号", "车牌号", "时段类型",
        "开始时刻", "结束时刻", "持续(分钟)",
        "电量(kWh)", "电价(元/kWh)", "费用(元)",
    ])
    period_map = {"peak": "峰时", "flat": "平时", "valley": "谷时"}
    for detail, bill_code, vehicle_id in rows:
        ws.append([
            detail.id, bill_code, vehicle_id,
            period_map.get(detail.period, detail.period),
            detail.start_time or "", detail.end_time or "",
            detail.duration_minutes or 0,
            round(detail.kwh or 0, 4),
            round(detail.rate or 0, 2),
            round(detail.fee or 0, 2),
        ])
    return _make_xlsx_response(wb, "bill_details.xlsx")
